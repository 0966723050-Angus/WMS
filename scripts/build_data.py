import json
import openpyxl

SRC = r"C:\Users\angus.ad\我的雲端硬碟\ATK\WMS.xlsx"
OUT = r"C:\ATK\Project\AI Agent\WMS-Web\data.json"


def norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def split_lines(v):
    if v is None:
        return []
    s = str(v)
    return [p.strip() for p in s.split("\n")]


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    ws1 = wb["在線製令一覽表"]
    ws2 = wb["在線製令物料明細"]

    projects = {}
    project_order = []

    for row in ws1.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code, name, order_no, prod_code, prod_name, spec, status, note = (
            row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
        )
        code = norm(code)
        if not code:
            continue
        if code not in projects:
            projects[code] = {"code": code, "name": norm(name), "orders": []}
            project_order.append(code)
        projects[code]["orders"].append({
            "orderNo": norm(order_no),
            "productCode": norm(prod_code),
            "productName": norm(prod_name),
            "spec": norm(spec),
            "status": norm(status),
            "note": norm(note),
        })

    materials = {}
    material_order = []

    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        (proj_code, order_no, prod_code, prod_name, seq, mat_code, mat_name, spec,
         mat_type, po_no, vendor, eta, need_qty, arrival_status, received_qty,
         issued_qty, remain_qty, stock_qty) = row[:18]

        order_no = norm(order_no)
        if order_no not in materials:
            materials[order_no] = []
            material_order.append(order_no)

        need_qty = need_qty if need_qty is not None else 0
        issued_qty = issued_qty if issued_qty is not None else 0
        stock_qty = stock_qty if stock_qty is not None else 0

        # 已到料 iff 需領用量 <= 庫存數量 + 已領用量(已入庫存或已實際領用的都算到料)
        arrival_status = "已到料" if need_qty <= stock_qty + issued_qty else "未到料"

        materials[order_no].append({
            "seq": seq if seq is not None else "",
            "materialCode": norm(mat_code),
            "materialName": norm(mat_name),
            "spec": norm(spec),
            "materialType": norm(mat_type),
            "poNos": split_lines(po_no),
            "vendors": split_lines(vendor),
            "etas": split_lines(eta),
            "needQty": need_qty,
            "arrivalStatus": arrival_status,
            "receivedQty": received_qty if received_qty is not None else 0,
            "issuedQty": issued_qty,
            "remainQty": remain_qty if remain_qty is not None else 0,
            "stockQty": stock_qty,
        })

    data = {
        "generatedFrom": "WMS.xlsx",
        "projects": [projects[c] for c in project_order],
        "materials": materials,
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    n_orders = sum(len(p["orders"]) for p in data["projects"])
    n_materials = sum(len(v) for v in materials.values())
    print(f"projects={len(data['projects'])} orders={n_orders} material_rows={n_materials}")


if __name__ == "__main__":
    main()
